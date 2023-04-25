#+*In[1]:*+
#[source, ipython3]
#----
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment 
from openpyxl.utils.cell import get_column_letter

class xlsx_dumper:
    
    ''' instantiate '''
    def __init__(self, view="new", fname="test.xlsx", template="wb_templates.xlsx"):
        self.view = view
        
        if fname == template:
            self.wbname = "test.xlsx"
        else:
            self.wbname = fname
         
        if self.view == "new": 
            self.template = template 
            self._init_workbook_new()
        else:
            self._init_workbook_old()
    
    ''' dump bmu/mmu data to xlsx '''
    def dump(self, ojson=None):
        
        if self.view == "new":
            return self._dump_new(ojson)
        else:    
            return self._dump_old(ojson)
   

    ''' save to xlsx'''
    def save(self, fname=None):
        if fname:
            self.wb.save(fname) 
        else:    
            self.wb.save(self.wbname)
    
    
    ''' get values of a column by the column name '''
    def getValueByColumn(self, chn=0, sheetNo=0, colname=None):
         
        ret = [] 
        if sheetNo == 0:
            ws = self.params[chn]['bmu']['ws']
            cells = ws.cell 
            num_of_col = ws.max_column
            num_of_row = ws.max_row
            for i in range(1,num_of_col): 
                if cells(row=4, column=i).value == colname: #  "Average Cell Voltage" :
                    for j in range(6, num_of_row):
                        ret.append(cells(row=j, column=i).value) 
                    return ret           
        else:
            ws = df.params[_channel]['mmus'][sheetNo-1]['ws']
            for i in range(1,num_of_col): 
                if cells(row=4, column=i).value == colname: #  "Average Cell Voltage" :
                    for j in range(5, num_of_row):
                        ret.append(cells(row=j, column=i).value) 
                    return ret

    ''' get column names '''
    def geColumnNames(self, chn=0, sheetNo=0):
        
        ret=[]
        if sheetNo == 0:
            ws = self.params[chn]['bmu']['ws']
            cells = ws.cell 
            num_of_col = ws.max_column
            for i in range(1,num_of_col): 
                if cells(row=5, column=i).value is None: 
                    ret.append(cells(row=4, column=i).value) 
                else:
                    ret.append(cells(row=5, column=i).value)            
        else:
            ws = df.params[chn]['mmus'][sheetNo-1]['ws']
            cells = ws.cell 
            num_of_col = ws.max_column
            for i in range(1,num_of_col):
                ret.append(cells(row=4, column=i).value)
              
        return ret
    
    ''' init workbook new format '''
    def _init_workbook_new(self):
        
        ''' initiate workbook from template ''' 
        self.wb = load_workbook(self.template)
          
        self.params=[{}, {}]
        
        chn0_params = self.params[0]
        chn1_params = self.params[1]
        
        ''' channel 0 params '''
        chn0_params["bmu"] = {"ws":self.wb["chn0-bmu"], "rows":{}} 
        chn0_params["mmus"] = []
        chn0_params["mmus"].append({"ws":self.wb["chn0-mmu0"], "rows":{}})
         
        ''' channel 1 params ''' 
        chn1wsbmu = self.wb.copy_worksheet(self.wb["chn0-bmu"])
        chn1wsbmu.title = "chn1-bmu"
        chn1wsmmu = self.wb.copy_worksheet(self.wb["chn0-mmu0"])
        chn1wsmmu.title="chn1-mmu0"
         
        chn1_params["bmu"] = {"ws": chn1wsbmu, "rows":{}} 
        chn1_params["mmus"] = []
        chn1_params["mmus"].append({"ws":chn1wsmmu, "rows":{}})
        
        self.bmu_cols = {}
        self.mmu_cols = {}
    
       
    ''' dump new format '''            
    def _dump_new(self, ojson=None):
        
        if ojson:
            
            chn = ojson['chn'] 
            if "mmu" in ojson.keys(): # ojson[mmu]:
                ws, ws_rows = self._getMMUParams(ojson['ft'], ojson['frame id'], self.params[chn]['mmus'])
            else:
                ws = self.params[chn]['bmu']['ws'] # self.ws0bmu
                ws_rows = self.params[chn]['bmu']['rows'] # self.row0bmu 
              
            name = ojson["name"]
            if name not in ws_rows.keys():
                if "mmu" in ojson.keys():
                    ws_rows[name] = 5
                else:
                    ws_rows[name] = 6
            else:
                ws_rows[name] += 1
                
            remarks = ojson["remark"]
            for byte in remarks:
                for i in byte[1:]:
                    field_name = i[1]  
                    if field_name == '':
                        field_name = i[3]
                    
                    col = self._getColumn(ws, field_name)
                    if col:
                        cell = ws.cell(row=ws_rows[name] ,column=col)
                        cell.value = i[2] 
                        
                        # print(i)
                    else:
                        if field_name != "~":
                            print("cannot find \'{}\'!".format(field_name))    

     
    ''' get column by frame name '''
    def _getColumn(self, ws, name):
        
        if "bmu" in ws.title:
            if name in self.bmu_cols.keys():
                return self.bmu_cols[name]
            else:
                cols = self.bmu_cols
        else:  # mmu
            if name in self.mmu_cols.keys():
                return self.mmu_cols[name]
            else:
                cols = self.mmu_cols
         
        ''' 
        for rows in ws["A3:BP3"]: # frame name at row 3 in template
            for cell in rows:
                if cell.value == name:
                    return cell.column
        '''
        
        for rows in ws["A4:CF4"]: # byte name at row 4 in template
            for cell in rows:
                if cell.value == name:
                    cols[name] = cell.column 
                    return cell.column
        for rows in ws["A5:CF5"]: # bit name at row 4 in template
            for cell in rows:
                if cell.value == name:
                    cols[name] = cell.column
                    return cell.column
      
    ''' get mmu params '''
    def _getMMUParams(self, ft, fid, mmus):
        
        mmu = fid & 0xff
        #toseq = (fid >> 8) & 0xff
        l = len(mmus)
        if mmu < l: 
            ws = mmus[mmu]['ws'] # get mmu ws  
        else:
            # create ws for mmu
            ws = mmus[-1]['ws'] # the last mmu 
            title = ws.title[:8] 
            
            while(l<=mmu):
                # print(ws.title)
                newws = self.wb.copy_worksheet(ws) # self.wb[ws]
                newws.title=title+str(l)
                mmus.append({"ws":newws, "rows":{}})
                ws = newws
                l += 1
             
          
        ws_rows = mmus[mmu]['rows']
        return ws, ws_rows
     
    
    ''' init workbook old format '''
    def _init_workbook_old(self):
    
        ''' create workbook '''
        self.wb= Workbook()
        
        self.ws0 = self.wb.create_sheet("chn0", 0)   # create sheet for chanel 0
        self.ws1 = self.wb.create_sheet("chn1", -1)  # create sheet for chanel 1

        ''' formatting '''
        self.ws0.column_dimensions['A'].width = 8
        self.ws0.column_dimensions['B'].width = 15
        self.ws0.column_dimensions['C'].width = 20
        self.ws0.column_dimensions['D'].width = 8
        self.ws0.column_dimensions['E'].width = 8
        self.ws0.column_dimensions['F'].width = 30
        self.ws0.column_dimensions['G'].width = 8
        self.ws0.column_dimensions['H'].width = 50
        self.ws0.column_dimensions['I'].width = 10

        self.ws1.column_dimensions['A'].width = 8
        self.ws1.column_dimensions['B'].width = 15
        self.ws1.column_dimensions['C'].width = 20
        self.ws1.column_dimensions['D'].width = 8
        self.ws1.column_dimensions['E'].width = 8
        self.ws1.column_dimensions['F'].width = 30
        self.ws1.column_dimensions['G'].width = 8
        self.ws1.column_dimensions['H'].width = 50
        self.ws1.column_dimensions['I'].width = 10

        ft = Font(bold=True, italic=True, name = 'Arial')
        garyFill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type = "solid")

        for r in self.ws0["A1:I1"]:
            i =0
            for c in r:
                c.font = ft
                c.fill = garyFill

        for r in self.ws1["A1:I1"]:
            i = 0
            for c in r:
                c.font = ft
                c.fill = garyFill 

        ''' set title '''
        self.ws0['A1'] = 'Intval'
        self.ws0['B1'] = 'ID'
        self.ws0['C1'] = 'Name'
        self.ws0['D1'] = 'Byte'
        self.ws0['E1'] = 'Bits'
        self.ws0['F1'] = 'Description'
        self.ws0['G1'] = 'Value'
        self.ws0['H1'] = 'Remark'
        self.ws0['I1'] = ''

        self.ws1['A1'] = 'Intval'
        self.ws1['B1'] = 'ID'
        self.ws1['C1'] = 'Name'
        self.ws1['D1'] = 'Byte'
        self.ws1['E1'] = 'Bits'
        self.ws1['F1'] = 'Description'
        self.ws1['G1'] = 'Value'
        self.ws1['H1'] = 'Remark'
        self.ws1['I1'] = ''

        ''' set position '''
        self.row0 = 2  # start row of sheet 'chn0' 
        self.row1 = 2  # start row of sheet 'chn1'
     
                 
    ''' dump old format '''
    def _dump_old(self, ojson=None):
        if ojson:
            if ojson['chn'] == 0:
                ws = self.ws0
                row = self.row0
            else:    
                ws = self.ws1
                row = self.row1

            ws['A'+str(row)] = ojson['intval'] 
            ws['B'+str(row)] = hex(ojson['frame id']) 
            ws['C'+str(row)] = ojson['name']

            rmk = ojson['remark']
            # row = 2
            for r in rmk: 
                col ='D' 
                ws[col+str(row)] = r[0]  # byte
                for i in range(1, len(r)):
                    col ='E'  
                    for j in range(0,len(r[i])):
                        ws[col+str(row)] = r[i][j]  # bits, description, value, remarks  
                        col = chr(ord(col)+1) 
                    row += 1

            if ojson['chn'] == 0:
                self.row0 = row
            else:    
                self.row1 = row

    
#----


#+*In[2]:*+
#[source, ipython3]
#----
if __name__ == "__main__":
    
    import ipynb
    from ipynb.fs.full.dataparser import *
    
    ft_list_1 = [
        FID_FT_INVALID ,
        FID_FT_BMU_SYSTEM_STATUS,    #     = 0xA0
        FID_FT_BMU_SYSTEM_MESSAGE_1, #     = 0x1A0
        FID_FT_BMU_SYSTEM_MESSAGE_2, #     = 0x2A0
        FID_FT_BMU_SYSTEM_MESSAGE_3, #     = 0x3A0
        FID_FT_BMU_INFORMATION_1, #        = 0x4A0
        FID_FT_BMU_INFORMATION_2, #        = 0x5A0
        FID_FT_BMU_INFORMATION_3, #        = 0x7A0
        FID_FT_BMU_STATISTIC_1, #          = 0x8A0
        FID_FT_BMU_STATISTIC_2, #          = 0x9A0
        FID_FT_BMU_VERSION_DATA, #         = 0xAA0
        FID_FT_BMU_CURRENT_INFO, #         = 0x11A0
        FID_FT_BMU_TOTAL_PRESSURE_COLLECTION, # = 0x14A0  
        
        FID_FT_MMU_SYSTEM_STATUS, # = 0x10000
        FID_FT_MMU_INFORMATION_1, # = 0x10200
        FID_FT_MMU_INFORMATION_2, # = 0x10300
        FID_FT_MMU_INFORMATION_3, # = 0x10400  
        FID_FT_MMU_CELL_VOL, #      = 0x20000
        FID_FT_MMU_CELL_TEMP, #     = 0x30000 
        FID_FT_MMU_VERSION_DATA #  = 0xA0000 
    ]

    count = 80000
 
    with open('data/00-01.txt', 'r') as f:
        lines = f.readlines()
        
    dumper = xlsx_dumper(view="new", fname="test.xlsx")
    # print(dumper.params)
    # dumper_old = xlsx_dumper(view="old", fname="test_old.xlsx")
       
    for line in lines:    
        if not line:
            break

        bframe = BMS_frame(line=line)
        if bframe.ft in ft_list_1 : # FID_FT_BMU_SYSTEM_STATUS:   
            ojson=bframe.dump_to_json()
            # print(ojson)
            dumper.dump(ojson=ojson)
            # dumper_old.dump(ojson=ojson)
    
        count -= 1
        if count == 0:
            break;
  
    dumper.save()
    # dumper_old.save()
        
#----


#+*In[ ]:*+
#[source, ipython3]
#----

#----
